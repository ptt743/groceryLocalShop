"""
App bán hàng local — Flask + PyWebView + SQLite (một file duy nhất).

Chạy thử (dev):
    pip install flask pywebview
    python app.py

Build ra .exe (chạy trên Windows):
    pip install pyinstaller
    pyinstaller --onefile --windowed --name "BanHang" app.py
    -> file nằm trong thư mục dist/

Dữ liệu lưu trong file shop.db, nằm CẠNH file app.py (hoặc cạnh .exe sau khi build).
Sao lưu = copy file shop.db. Toàn bộ chạy offline, không cần internet.
"""

import os
import sys
import sqlite3
import threading
from flask import Flask, request, jsonify, render_template_string, session
from werkzeug.security import generate_password_hash, check_password_hash

# ---------------------------------------------------------------------------
# Đường dẫn DB: đặt cạnh app.py khi dev, cạnh .exe khi đã đóng gói.
# ---------------------------------------------------------------------------
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.environ.get("DB_PATH") or os.path.join(BASE_DIR, "shop.db")

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Cơ sở dữ liệu
# ---------------------------------------------------------------------------
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS products (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            name    TEXT    NOT NULL,
            barcode TEXT,
            price   INTEGER NOT NULL DEFAULT 0,
            stock   INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS sales (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            total     INTEGER NOT NULL,
            paid      INTEGER NOT NULL,
            change    INTEGER NOT NULL,
            created_at TEXT   NOT NULL DEFAULT (datetime('now','+7 hours'))
        );

        CREATE TABLE IF NOT EXISTS sale_items (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id    INTEGER NOT NULL,
            product_id INTEGER,
            name       TEXT    NOT NULL,
            price      INTEGER NOT NULL,
            qty        INTEGER NOT NULL
        );

        CREATE TABLE IF NOT EXISTS debt_payments (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            phone      TEXT    NOT NULL,
            amount     INTEGER NOT NULL,
            created_at TEXT    NOT NULL DEFAULT (datetime('now','+7 hours'))
        );

        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        );
        """
    )
    # migration: thêm cột debtor_phone vào bảng sales nếu DB cũ chưa có
    cols = [r[1] for r in conn.execute("PRAGMA table_info(sales)").fetchall()]
    if "debtor_phone" not in cols:
        conn.execute("ALTER TABLE sales ADD COLUMN debtor_phone TEXT")

    # mật khẩu mặc định "123456" (lưu dạng băm) nếu chưa có
    row = conn.execute("SELECT value FROM settings WHERE key='password_hash'").fetchone()
    if not row:
        conn.execute(
            "INSERT INTO settings (key, value) VALUES ('password_hash', ?)",
            (generate_password_hash("123456"),),
        )
    # khoá ký session, ổn định qua các lần khởi động lại
    row = conn.execute("SELECT value FROM settings WHERE key='secret_key'").fetchone()
    if not row:
        conn.execute(
            "INSERT INTO settings (key, value) VALUES ('secret_key', ?)",
            (os.urandom(24).hex(),),
        )
    conn.commit()
    conn.close()


def get_setting(key):
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row["value"] if row else None


def set_setting(key, value):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?,?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# API sản phẩm
# ---------------------------------------------------------------------------
@app.route("/api/products", methods=["GET"])
def list_products():
    q = (request.args.get("q") or "").strip()
    conn = get_db()
    if q:
        like = f"%{q}%"
        rows = conn.execute(
            "SELECT * FROM products WHERE name LIKE ? OR barcode LIKE ? ORDER BY name",
            (like, like),
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM products ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/products/barcode/<code>", methods=["GET"])
def find_by_barcode(code):
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM products WHERE barcode = ?", (code.strip(),)
    ).fetchone()
    conn.close()
    if row:
        return jsonify(dict(row))
    return jsonify({"error": "not_found"}), 404


@app.route("/api/lookup/<code>", methods=["GET"])
def lookup_openfoodfacts(code):
    """Tra tên sản phẩm theo mã vạch trên Open Food Facts (nguồn mở).
    Gọi từ server để gắn được User-Agent theo yêu cầu của OFF.
    Luôn trả 200; nếu không có mạng / không tìm thấy thì found=false."""
    import urllib.request
    import json as _json

    code = code.strip()
    url = (
        "https://world.openfoodfacts.org/api/v2/product/"
        f"{code}.json?fields=product_name,product_name_vi,brands,quantity"
    )
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "BanHangLocal/1.0 (pos-local-app)"},
    )
    try:
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
    except Exception:
        return jsonify({"found": False})

    if data.get("status") == 1:
        p = data.get("product", {}) or {}
        name = (p.get("product_name_vi") or p.get("product_name") or "").strip()
        qty = (p.get("quantity") or "").strip()
        full = name
        if qty and qty.lower() not in name.lower():
            full = (name + " " + qty).strip()
        return jsonify({
            "found": bool(name),
            "name": full,
            "brand": (p.get("brands") or "").strip(),
        })
    return jsonify({"found": False})


@app.route("/api/products", methods=["POST"])
def add_product():
    d = request.get_json(force=True)
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Tên sản phẩm không được để trống"}), 400
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO products (name, barcode, price, stock) VALUES (?,?,?,?)",
        (
            name,
            (d.get("barcode") or "").strip() or None,
            int(d.get("price") or 0),
            int(d.get("stock") or 0),
        ),
    )
    conn.commit()
    pid = cur.lastrowid
    row = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


@app.route("/api/products/<int:pid>", methods=["PUT"])
def update_product(pid):
    d = request.get_json(force=True)
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Tên sản phẩm không được để trống"}), 400
    conn = get_db()
    conn.execute(
        "UPDATE products SET name=?, barcode=?, price=?, stock=? WHERE id=?",
        (
            name,
            (d.get("barcode") or "").strip() or None,
            int(d.get("price") or 0),
            int(d.get("stock") or 0),
            pid,
        ),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
    conn.close()
    return jsonify(dict(row))


@app.route("/api/products/<int:pid>/price", methods=["POST"])
def update_price(pid):
    d = request.get_json(force=True)
    price = int(d.get("price") or 0)
    if price <= 0:
        return jsonify({"error": "Giá không hợp lệ"}), 400
    conn = get_db()
    conn.execute("UPDATE products SET price=? WHERE id=?", (price, pid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/products/<int:pid>", methods=["DELETE"])
def delete_product(pid):
    conn = get_db()
    conn.execute("DELETE FROM products WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# API thanh toán
# ---------------------------------------------------------------------------
@app.route("/api/checkout", methods=["POST"])
def checkout():
    d = request.get_json(force=True)
    items = d.get("items") or []
    if not items:
        return jsonify({"error": "Giỏ hàng trống"}), 400

    total = sum(int(i["price"]) * int(i["qty"]) for i in items)

    # nếu có SĐT người nợ -> đây là đơn ghi nợ (chưa trả tiền)
    phone = (d.get("debtor_phone") or "").strip()
    if phone:
        paid = 0
        change = 0
    else:
        paid = int(d.get("paid") or 0)
        change = paid - total

    conn = get_db()
    cur = conn.execute(
        "INSERT INTO sales (total, paid, change, debtor_phone, created_at) "
        "VALUES (?,?,?,?, datetime('now','+7 hours'))",
        (total, paid, max(change, 0), phone or None),
    )
    sale_id = cur.lastrowid
    for i in items:
        conn.execute(
            "INSERT INTO sale_items (sale_id, product_id, name, price, qty) VALUES (?,?,?,?,?)",
            (sale_id, i.get("id"), i["name"], int(i["price"]), int(i["qty"])),
        )
        # trừ tồn kho nếu là sản phẩm có trong DB
        if i.get("id"):
            conn.execute(
                "UPDATE products SET stock = stock - ? WHERE id = ?",
                (int(i["qty"]), i["id"]),
            )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "sale_id": sale_id, "total": total,
                    "change": change, "is_debt": bool(phone)})


# ---------------------------------------------------------------------------
# API lịch sử thanh toán
# ---------------------------------------------------------------------------
@app.route("/api/sales", methods=["GET"])
def list_sales():
    conn = get_db()

    # ngày đang xem (mặc định: hôm nay)
    day = (request.args.get("date") or "").strip()
    if not day:
        row = conn.execute("SELECT date('now','+7 hours') AS d").fetchone()
        day = row["d"]

    sales = conn.execute(
        "SELECT * FROM sales WHERE date(created_at)=? ORDER BY id DESC",
        (day,),
    ).fetchall()
    result = []
    for s in sales:
        items = conn.execute(
            "SELECT name, price, qty FROM sale_items WHERE sale_id=?", (s["id"],)
        ).fetchall()
        d = dict(s)
        d["items"] = [dict(i) for i in items]
        result.append(d)

    # doanh thu & số đơn của ngày đang xem
    stat = conn.execute(
        "SELECT COALESCE(SUM(total),0) AS revenue, COUNT(*) AS count "
        "FROM sales WHERE date(created_at)=?",
        (day,),
    ).fetchone()

    # danh sách các ngày có giao dịch (mới nhất trước) để chọn
    dates = conn.execute(
        "SELECT date(created_at) AS d FROM sales GROUP BY d ORDER BY d DESC"
    ).fetchall()

    conn.close()
    return jsonify({
        "date": day,
        "sales": result,
        "day_revenue": stat["revenue"],
        "day_count": stat["count"],
        "dates": [r["d"] for r in dates],
    })


@app.route("/api/revenue/yearly", methods=["GET"])
def revenue_yearly():
    conn = get_db()

    # năm đang xem (mặc định: năm hiện tại)
    year = (request.args.get("year") or "").strip()
    if not year:
        year = conn.execute(
            "SELECT strftime('%Y','now','+7 hours') AS y"
        ).fetchone()["y"]

    # doanh thu & số đơn từng tháng trong năm
    rows = conn.execute(
        "SELECT strftime('%m', created_at) AS m, "
        "COALESCE(SUM(total),0) AS revenue, COUNT(*) AS count "
        "FROM sales WHERE strftime('%Y', created_at)=? "
        "GROUP BY m",
        (year,),
    ).fetchall()
    by_month = {r["m"]: r for r in rows}
    months = []
    for i in range(1, 13):
        key = f"{i:02d}"
        r = by_month.get(key)
        months.append({
            "month": i,
            "revenue": r["revenue"] if r else 0,
            "count": r["count"] if r else 0,
        })

    total = conn.execute(
        "SELECT COALESCE(SUM(total),0) AS t, COUNT(*) AS c "
        "FROM sales WHERE strftime('%Y', created_at)=?",
        (year,),
    ).fetchone()

    # danh sách các năm có giao dịch
    years = conn.execute(
        "SELECT strftime('%Y', created_at) AS y FROM sales GROUP BY y ORDER BY y DESC"
    ).fetchall()

    conn.close()
    return jsonify({
        "year": year,
        "total": total["t"],
        "count": total["c"],
        "months": months,
        "years": [r["y"] for r in years],
    })


@app.route("/api/export", methods=["GET"])
def export_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from flask import send_file

    head_fill = PatternFill("solid", fgColor="0F766E")

    # ----- Xuất báo cáo theo NĂM: chi tiết + theo sản phẩm + theo tháng -----
    year = (request.args.get("year") or "").strip()
    if year:
        conn = get_db()
        # chi tiết từng món của từng đơn trong năm
        detail = conn.execute(
            "SELECT s.id, s.created_at, i.name, p.barcode AS barcode, i.price, i.qty, s.total, s.debtor_phone "
            "FROM sales s JOIN sale_items i ON i.sale_id=s.id "
            "LEFT JOIN products p ON p.id = i.product_id "
            "WHERE strftime('%Y', s.created_at)=? ORDER BY s.id, i.id",
            (year,),
        ).fetchall()
        # tổng hợp theo sản phẩm
        by_product = conn.execute(
            "SELECT i.name AS name, MAX(p.barcode) AS barcode, "
            "SUM(i.qty) AS qty, SUM(i.price*i.qty) AS amount "
            "FROM sales s JOIN sale_items i ON i.sale_id=s.id "
            "LEFT JOIN products p ON p.id = i.product_id "
            "WHERE strftime('%Y', s.created_at)=? GROUP BY i.name ORDER BY amount DESC",
            (year,),
        ).fetchall()
        # tổng hợp theo tháng
        rows = conn.execute(
            "SELECT strftime('%m', created_at) AS m, "
            "COALESCE(SUM(total),0) AS revenue, COUNT(*) AS count "
            "FROM sales WHERE strftime('%Y', created_at)=? GROUP BY m",
            (year,),
        ).fetchall()
        conn.close()
        by_month = {r["m"]: r for r in rows}

        def style_header(ws):
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = head_fill
                c.alignment = Alignment(horizontal="center")

        wb = Workbook()

        # ===== Sheet 1: Chi tiết =====
        ws = wb.active
        ws.title = "Chi tiết"
        ws.append(["Mã đơn", "Thời gian", "Sản phẩm", "Mã vạch", "Đơn giá", "Số lượng", "Thành tiền", "Loại"])
        style_header(ws)
        grand_total = 0
        seen = set()
        for r in detail:
            line = r["price"] * r["qty"]
            kind = "Nợ" if r["debtor_phone"] else "Bán"
            ws.append([r["id"], r["created_at"], r["name"], r["barcode"] or "",
                       r["price"], r["qty"], line, kind])
            if r["id"] not in seen:
                grand_total += r["total"]
                seen.add(r["id"])
        ws.append([])
        ws.append(["", "", "", "", "", "Tổng doanh thu", grand_total, ""])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        for col, w in zip("ABCDEFGH", [10, 20, 32, 16, 14, 10, 16, 8]):
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
            for c in row:
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0"

        # ===== Sheet 2: Theo sản phẩm =====
        ws2 = wb.create_sheet("Theo sản phẩm")
        ws2.append(["Sản phẩm", "Mã vạch", "Số lượng bán", "Doanh thu"])
        style_header(ws2)
        for r in by_product:
            ws2.append([r["name"], r["barcode"] or "", r["qty"], r["amount"]])
        for col, w in zip("ABCD", [32, 16, 14, 18]):
            ws2.column_dimensions[col].width = w
        for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
            for c in row:
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0"

        # ===== Sheet 3: Theo tháng =====
        ws3 = wb.create_sheet("Theo tháng")
        ws3.append(["Tháng", "Số đơn", "Doanh thu"])
        style_header(ws3)
        year_total = 0
        for i in range(1, 13):
            r = by_month.get(f"{i:02d}")
            rev = r["revenue"] if r else 0
            cnt = r["count"] if r else 0
            year_total += rev
            ws3.append([f"Tháng {i}", cnt, rev])
        ws3.append([])
        ws3.append(["Tổng cả năm", "", year_total])
        for c in ws3[ws3.max_row]:
            c.font = Font(bold=True)
        for col, w in zip("ABC", [16, 12, 18]):
            ws3.column_dimensions[col].width = w
        for row in ws3.iter_rows(min_row=2, min_col=3, max_col=3):
            for c in row:
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"doanh-thu-nam-{year}.xlsx",
        )

    day = (request.args.get("date") or "").strip()  # rỗng hoặc "all" = tất cả

    conn = get_db()
    if day and day != "all":
        rows = conn.execute(
            "SELECT s.id, s.created_at, i.name, p.barcode AS barcode, i.price, i.qty, s.total, s.debtor_phone "
            "FROM sales s JOIN sale_items i ON i.sale_id=s.id "
            "LEFT JOIN products p ON p.id = i.product_id "
            "WHERE date(s.created_at)=? ORDER BY s.id, i.id",
            (day,),
        ).fetchall()
        fname = f"doanh-thu-{day}.xlsx"
    else:
        rows = conn.execute(
            "SELECT s.id, s.created_at, i.name, p.barcode AS barcode, i.price, i.qty, s.total, s.debtor_phone "
            "FROM sales s JOIN sale_items i ON i.sale_id=s.id "
            "LEFT JOIN products p ON p.id = i.product_id "
            "ORDER BY s.id, i.id"
        ).fetchall()
        fname = "doanh-thu-tat-ca.xlsx"
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Doanh thu"

    headers = ["Mã đơn", "Thời gian", "Sản phẩm", "Mã vạch", "Đơn giá", "Số lượng", "Thành tiền", "Loại"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="0F766E")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    grand_total = 0
    seen_orders = set()
    for r in rows:
        line_total = r["price"] * r["qty"]
        kind = "Nợ" if r["debtor_phone"] else "Bán"
        ws.append([r["id"], r["created_at"], r["name"], r["barcode"] or "",
                   r["price"], r["qty"], line_total, kind])
        if r["id"] not in seen_orders:
            grand_total += r["total"]
            seen_orders.add(r["id"])

    # dòng tổng
    ws.append([])
    ws.append(["", "", "", "", "", "Tổng doanh thu", grand_total, ""])
    last = ws.max_row
    for c in ws[last]:
        c.font = Font(bold=True)

    # định dạng cột tiền + độ rộng
    widths = [10, 20, 30, 16, 14, 10, 16, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


# ---------------------------------------------------------------------------
# API sổ nợ
# ---------------------------------------------------------------------------
@app.route("/api/debtors", methods=["GET"])
def list_debtors():
    conn = get_db()
    rows = conn.execute(
        "SELECT s.debtor_phone AS phone, "
        "COALESCE(SUM(s.total),0) - "
        "COALESCE((SELECT SUM(amount) FROM debt_payments p WHERE p.phone = s.debtor_phone),0) "
        "AS remaining "
        "FROM sales s WHERE s.debtor_phone IS NOT NULL "
        "GROUP BY s.debtor_phone HAVING remaining > 0 "
        "ORDER BY remaining DESC LIMIT 50"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/debt/<phone>", methods=["GET"])
def get_debt(phone):
    phone = phone.strip()
    conn = get_db()

    # các đơn đã ghi nợ cho SĐT này
    sales = conn.execute(
        "SELECT * FROM sales WHERE debtor_phone=? ORDER BY id DESC", (phone,)
    ).fetchall()
    orders = []
    for s in sales:
        items = conn.execute(
            "SELECT name, price, qty FROM sale_items WHERE sale_id=?", (s["id"],)
        ).fetchall()
        d = dict(s)
        d["items"] = [dict(i) for i in items]
        orders.append(d)

    debt_total = conn.execute(
        "SELECT COALESCE(SUM(total),0) AS t FROM sales WHERE debtor_phone=?", (phone,)
    ).fetchone()["t"]
    paid_total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) AS t FROM debt_payments WHERE phone=?", (phone,)
    ).fetchone()["t"]
    payments = conn.execute(
        "SELECT amount, created_at FROM debt_payments WHERE phone=? ORDER BY id DESC",
        (phone,),
    ).fetchall()
    conn.close()

    return jsonify({
        "phone": phone,
        "debt_total": debt_total,        # tổng tiền đã mua nợ
        "paid_total": paid_total,        # tổng đã trả
        "remaining": debt_total - paid_total,  # còn nợ
        "orders": orders,
        "payments": [dict(p) for p in payments],
    })


@app.route("/api/debt/<phone>/pay", methods=["POST"])
def pay_debt(phone):
    phone = phone.strip()
    d = request.get_json(force=True)
    amount = int(d.get("amount") or 0)
    if amount <= 0:
        return jsonify({"error": "Số tiền trả phải lớn hơn 0"}), 400

    conn = get_db()
    conn.execute(
        "INSERT INTO debt_payments (phone, amount, created_at) "
        "VALUES (?,?, datetime('now','+7 hours'))",
        (phone, amount),
    )
    conn.commit()
    debt_total = conn.execute(
        "SELECT COALESCE(SUM(total),0) AS t FROM sales WHERE debtor_phone=?", (phone,)
    ).fetchone()["t"]
    paid_total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) AS t FROM debt_payments WHERE phone=?", (phone,)
    ).fetchone()["t"]
    conn.close()
    return jsonify({"ok": True, "remaining": debt_total - paid_total})


# ---------------------------------------------------------------------------
# Giao diện (nhúng trực tiếp)
# ---------------------------------------------------------------------------
PAGE = r"""
<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Bán hàng</title>
<style>
  :root{
    --bg:#ffffff;
    --surface:#f6f7f8;
    --line:#e6e8eb;
    --ink:#15171a;
    --muted:#6b7177;
    --accent:#0f766e;        /* xanh ngọc đậm */
    --accent-soft:#e6f4f2;
    --danger:#b4232a;
    --radius:10px;
  }
  *{box-sizing:border-box}
  html,body{margin:0;height:100%}
  body{
    font-family:system-ui,-apple-system,"Segoe UI",Roboto,sans-serif;
    color:var(--ink);background:var(--bg);
    font-size:15px;line-height:1.45;
  }
  .num{font-variant-numeric:tabular-nums;font-feature-settings:"tnum"}

  /* khung */
  header{
    display:flex;align-items:center;gap:24px;
    padding:0 22px;height:56px;border-bottom:1px solid var(--line);
  }
  header .brand{font-weight:650;letter-spacing:-.01em}
  nav{display:flex;gap:4px;margin-left:8px}
  nav button{
    border:0;background:none;color:var(--muted);
    padding:8px 14px;border-radius:8px;cursor:pointer;font-size:15px;
  }
  nav button:hover{background:var(--surface)}
  nav button.active{background:var(--ink);color:#fff}

  .header-right{margin-left:auto;display:flex;gap:4px}
  .link-btn{border:0;background:none;color:var(--muted);padding:8px 12px;border-radius:8px;cursor:pointer;font-size:14px;font-family:inherit}
  .link-btn:hover{background:var(--surface);color:var(--ink)}

  /* màn hình đăng nhập */
  .login-screen{position:fixed;inset:0;background:var(--bg);display:none;
    align-items:center;justify-content:center;z-index:100}
  .login-screen.show{display:flex}
  .login-card{width:100%;max-width:320px;text-align:center;padding:24px}
  .login-card h1{font-size:26px;margin:0 0 4px;letter-spacing:-.02em}
  .login-sub{color:var(--muted);margin:0 0 22px}
  .login-card input{width:100%;padding:12px 14px;border:1px solid var(--line);border-radius:8px;
    font-size:16px;outline:none;margin-bottom:12px;font-family:inherit;text-align:center}
  .login-card input:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .login-card .btn{width:100%}

  main{padding:22px;max-width:1180px;margin:0 auto}
  .hide{display:none}

  /* tổng quát */
  h2{font-size:14px;font-weight:600;color:var(--muted);
     text-transform:uppercase;letter-spacing:.04em;margin:0 0 12px}
  input,button{font-family:inherit;font-size:15px}
  .field input{
    width:100%;padding:10px 12px;border:1px solid var(--line);
    border-radius:8px;background:#fff;outline:none;
  }
  .field input:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .field label{display:block;font-size:13px;color:var(--muted);margin:0 0 5px}
  .btn{
    border:0;border-radius:8px;padding:10px 16px;cursor:pointer;font-weight:550;
    background:var(--accent);color:#fff;
  }
  .btn:hover{filter:brightness(.95)}
  .btn.ghost{background:var(--surface);color:var(--ink)}
  .btn.ghost:hover{background:var(--line)}
  .btn:disabled{opacity:.4;cursor:not-allowed}

  /* ====== BÁN HÀNG ====== */
  .sale{display:grid;grid-template-columns:1fr 380px;gap:22px;align-items:start}
  .scan{display:flex;gap:10px;margin-bottom:16px}
  .scan input{flex:1}
  .plist{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:10px}
  .pcard{
    border:1px solid var(--line);border-radius:var(--radius);padding:12px;
    cursor:pointer;background:#fff;text-align:left;transition:.12s;
  }
  .pcard:hover{border-color:var(--accent);background:var(--accent-soft)}
  .pcard .pn{font-weight:550;margin-bottom:4px}
  .pcard .pp{color:var(--accent);font-weight:600}
  .pcard .ps{font-size:12px;color:var(--muted);margin-top:2px}

  .cart{
    border:1px solid var(--line);border-radius:var(--radius);
    background:#fff;position:sticky;top:22px;overflow:hidden;
  }
  .cart .ch{padding:14px 16px;border-bottom:1px solid var(--line);font-weight:600}
  .citems{max-height:320px;overflow:auto}
  .citem{display:flex;align-items:center;gap:8px;padding:10px 16px;border-bottom:1px solid var(--surface)}
  .citem .ci-n{flex:1;min-width:0}
  .citem .ci-n div:first-child{font-weight:500;font-size:13px;color:var(--muted);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .citem .ci-n div:last-child{font-size:12px;color:var(--muted)}
  .ci-price{cursor:pointer;color:var(--accent)!important;font-size:19px!important;font-weight:700!important;display:inline-block;border-radius:5px;padding:2px 5px;margin-top:3px}
  .ci-price:hover{background:var(--accent-soft)}
  .qty{display:flex;align-items:center;gap:6px}
  .qty button{width:26px;height:26px;border:1px solid var(--line);background:#fff;border-radius:6px;cursor:pointer}
  .qty span{min-width:20px;text-align:center}
  .ci-rm{border:0;background:none;color:var(--danger);cursor:pointer;font-size:18px;line-height:1;padding:0 2px}
  .cempty{padding:40px 16px;text-align:center;color:var(--muted)}
  .csum{padding:14px 16px;border-top:1px solid var(--line)}
  .row{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px}
  .row .lbl{color:var(--muted)}
  .total .lbl{font-weight:600;color:var(--ink)}
  .total .val{font-size:22px;font-weight:700;color:var(--accent)}
  .change .val{font-weight:600}
  .pay-input{width:150px;text-align:right}

  /* ====== SẢN PHẨM ====== */
  .prodgrid{display:grid;grid-template-columns:320px 1fr;gap:22px;align-items:start}
  .form-card,.table-card{border:1px solid var(--line);border-radius:var(--radius);background:#fff;padding:18px}
  .form-card .field{margin-bottom:12px}
  .form-actions{display:flex;gap:8px;margin-top:4px}
  .search{margin-bottom:14px}
  table{width:100%;border-collapse:collapse}
  th,td{text-align:left;padding:10px 12px;border-bottom:1px solid var(--surface)}
  th{font-size:12px;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);font-weight:600}
  td.r,th.r{text-align:right}
  .t-actions button{border:0;background:none;cursor:pointer;color:var(--muted);padding:4px 6px;border-radius:6px}
  .t-actions button:hover{background:var(--surface);color:var(--ink)}
  .t-actions .del:hover{color:var(--danger)}
  .empty-row td{text-align:center;color:var(--muted);padding:30px}

  /* ====== LỊCH SỬ ====== */
  .year-block{border:1px solid var(--line);border-radius:var(--radius);background:#fff;padding:18px;margin-bottom:8px}
  .year-head{display:flex;gap:14px;align-items:flex-end;flex-wrap:wrap;margin-bottom:18px}
  .month-grid{display:grid;grid-template-columns:repeat(12,1fr);gap:8px}
  .month-cell{text-align:center}
  .month-name{font-size:12px;color:var(--muted);margin-bottom:6px}
  .month-bar{height:90px;background:var(--surface);border-radius:6px;display:flex;align-items:flex-end;overflow:hidden}
  .month-fill{width:100%;background:var(--accent);border-radius:6px 6px 0 0;transition:height .3s;min-height:2px}
  .month-rev{font-size:11px;margin-top:6px;font-weight:600}
  .month-cnt{font-size:11px;color:var(--muted)}
  @media(max-width:760px){
    .month-grid{grid-template-columns:repeat(6,1fr)}
    .month-rev{font-size:10px}
  }
  .hist-top{display:flex;gap:14px;align-items:flex-end;margin-bottom:18px;flex-wrap:wrap}
  .hist-pick label{display:block;font-size:13px;color:var(--muted);margin-bottom:4px}
  .hist-pick select{padding:10px 12px;border:1px solid var(--line);border-radius:8px;background:#fff;
    font-family:inherit;font-size:15px;outline:none;min-width:160px}
  .hist-pick select:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .hist-actions{margin-left:auto;display:flex;gap:8px}
  .stat{border:1px solid var(--line);border-radius:var(--radius);background:#fff;padding:14px 18px;min-width:180px}
  .stat-lbl{font-size:13px;color:var(--muted);margin-bottom:4px}
  .stat-val{font-size:24px;font-weight:700;color:var(--accent)}
  .sales-list{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:14px}
  .bill{border:1px solid var(--line);border-radius:var(--radius);background:#fff;overflow:hidden}
  .bill-head{display:flex;justify-content:space-between;align-items:center;padding:10px 14px;
    border-bottom:1px solid var(--surface);background:var(--surface)}
  .bill-id{font-weight:600}
  .bill-time{font-size:12px;color:var(--muted)}
  .bill-body{padding:8px 14px}
  .bill-item{display:flex;justify-content:space-between;padding:3px 0;font-size:14px}
  .bill-foot{padding:10px 14px;border-top:1px solid var(--surface)}
  .bill-foot .lbl{color:var(--muted)}
  .bill-total{font-weight:700;color:var(--accent)}
  .bill-sub{font-size:13px;margin-top:4px}

  /* ====== SỔ NỢ ====== */
  .btn-debt{background:#b45309}            /* cam đất, phân biệt với nút thanh toán */
  .btn-red{background:var(--danger)}
  .debt-search{display:flex;gap:10px;margin-bottom:18px;max-width:520px}
  .debt-search input{flex:1;padding:10px 12px;border:1px solid var(--line);border-radius:8px;outline:none;font-size:15px;font-family:inherit}
  .debt-search input:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .debt-summary{display:flex;justify-content:space-between;gap:20px;align-items:flex-end;flex-wrap:wrap;
    border:1px solid var(--line);border-radius:var(--radius);background:#fff;padding:18px;margin-bottom:18px}
  .debt-pay{display:flex;gap:8px;align-items:center}
  .debt-pay input{padding:10px 12px;border:1px solid var(--line);border-radius:8px;outline:none;
    font-size:15px;font-family:inherit;text-align:right;width:180px}
  .debt-pay input:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .debt-cols{display:grid;grid-template-columns:1fr 1fr;gap:20px;align-items:start}
  .pay-history{border:1px solid var(--line);border-radius:var(--radius);background:#fff;padding:10px 14px}
  .pay-history .bill-item{border-bottom:1px solid var(--surface)}
  @media(max-width:760px){.debt-cols{grid-template-columns:1fr}}
  .debtor-list{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:10px;margin-bottom:18px}
  .debtor-row{display:flex;align-items:center;gap:12px;border:1px solid var(--line);border-radius:var(--radius);
    background:#fff;padding:12px 14px;cursor:pointer;text-align:left;font-family:inherit;font-size:15px;transition:.12s}
  .debtor-row:hover{border-color:var(--danger);background:#fdf2f2}
  .debtor-rank{width:24px;height:24px;flex:none;border-radius:50%;background:var(--surface);color:var(--muted);
    display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:600}
  .debtor-phone{flex:1;font-weight:550}
  .debtor-amt{color:var(--danger);font-weight:700}
  #view-debt .bill{margin-bottom:12px}
  #view-debt .pay-history .bill-item{padding:10px 2px}

  /* toast */
  #toast{position:fixed;bottom:24px;left:50%;transform:translateX(-50%) translateY(20px);
    background:var(--ink);color:#fff;padding:11px 18px;border-radius:8px;opacity:0;
    transition:.2s;pointer-events:none;font-size:14px;z-index:50}
  #toast.show{opacity:1;transform:translateX(-50%) translateY(0)}

  /* hộp thoại thêm sản phẩm mới */
  .modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;
    align-items:center;justify-content:center;z-index:60;padding:20px}
  .modal-overlay.show{display:flex}
  .modal{background:#fff;border-radius:var(--radius);padding:22px;width:100%;max-width:380px;
    box-shadow:0 12px 48px rgba(0,0,0,.25)}
  .modal h2{margin-top:0}
  .np-code-row{color:var(--muted);font-size:14px;margin-bottom:14px}
  .modal .field{margin-bottom:12px}
  .modal-actions{display:flex;gap:8px;margin-top:6px}
  .modal-actions .btn{flex:1}
</style>
</head>
<body>
<header>
  <span class="brand">Bán hàng</span>
  <nav>
    <button id="tab-sale" class="active" onclick="showTab('sale')">Bán hàng</button>
    <button id="tab-prod" onclick="showTab('prod')">Sản phẩm</button>
    <button id="tab-hist" onclick="showTab('hist')">Lịch sử</button>
    <button id="tab-debt" onclick="showTab('debt')">Sổ nợ</button>
  </nav>
  <div class="header-right">
    <button class="link-btn" onclick="openChangePw()">Đổi mật khẩu</button>
    <button class="link-btn" onclick="doLogout()">Đăng xuất</button>
  </div>
</header>

<main>
  <!-- ================= BÁN HÀNG ================= -->
  <section id="view-sale">
    <div class="sale">
      <div>
        <div class="scan">
          <input id="scan" placeholder="Quét mã vạch — tự thêm vào giỏ…" autocomplete="off">
          <button class="btn ghost" onclick="document.getElementById('scan').focus()">Quét</button>
        </div>
        <h2>Chọn sản phẩm</h2>
        <div id="saleList" class="plist"></div>
      </div>

      <aside class="cart">
        <div class="ch">Giỏ hàng</div>
        <div id="cartItems" class="citems"></div>
        <div class="csum">
          <div class="row total"><span class="lbl">Tổng cộng</span><span id="total" class="val num">0 ₫</span></div>
          <div class="row"><span class="lbl">Khách đưa</span>
            <input id="paid" class="field-inline pay-input num" inputmode="numeric" placeholder="0" oninput="onPaid()"
                   style="padding:7px 10px;border:1px solid var(--line);border-radius:8px;outline:none">
          </div>
          <div class="row change"><span class="lbl">Tiền thối</span><span id="change" class="val num">0 ₫</span></div>
          <button id="payBtn" class="btn" style="width:100%;margin-top:6px" onclick="checkout()" disabled>Thanh toán</button>
          <button id="debtBtn" class="btn btn-debt" style="width:100%;margin-top:8px" onclick="checkoutDebt()" disabled>Ghi nợ</button>
          <button class="btn ghost" style="width:100%;margin-top:8px" onclick="clearCart()">Xoá giỏ</button>
        </div>
      </aside>
    </div>
  </section>

  <!-- ================= SẢN PHẨM ================= -->
  <section id="view-prod" class="hide">
    <div class="prodgrid">
      <div class="form-card">
        <h2 id="formTitle">Thêm sản phẩm</h2>
        <input type="hidden" id="pid">
        <div class="field"><label>Tên sản phẩm</label><input id="f-name" autocomplete="off"></div>
        <div class="field"><label>Mã vạch / SKU</label><input id="f-barcode" placeholder="Quét hoặc nhập mã…" autocomplete="off"></div>
        <div class="field"><label>Giá bán (₫)</label><input id="f-price" inputmode="numeric" autocomplete="off"></div>
        <div class="field"><label>Tồn kho</label><input id="f-stock" inputmode="numeric" autocomplete="off"></div>
        <div class="form-actions">
          <button class="btn" onclick="saveProduct()">Lưu</button>
          <button class="btn ghost" id="cancelEdit" onclick="resetForm()" style="display:none">Huỷ</button>
        </div>
      </div>

      <div class="table-card">
        <div class="search field"><input id="search" placeholder="Tìm theo tên hoặc mã vạch…" oninput="loadProducts(this.value)" autocomplete="off"></div>
        <table>
          <thead><tr>
            <th>Tên</th><th>Mã vạch</th><th class="r">Giá</th><th class="r">Tồn</th><th></th>
          </tr></thead>
          <tbody id="prodTable"></tbody>
        </table>
      </div>
    </div>
  </section>

  <!-- ================= LỊCH SỬ ================= -->
  <section id="view-hist" class="hide">

    <!-- Doanh thu theo năm -->
    <div class="year-block">
      <div class="year-head">
        <div class="hist-pick">
          <label>Doanh thu năm</label>
          <select id="yearPick" onchange="loadYearly(this.value)"></select>
        </div>
        <div class="stat">
          <div class="stat-lbl">Tổng doanh thu năm</div>
          <div class="stat-val num" id="yearTotal">0 ₫</div>
        </div>
        <div class="stat">
          <div class="stat-lbl">Số đơn cả năm</div>
          <div class="stat-val num" id="yearCount">0</div>
        </div>
        <button class="btn" style="margin-left:auto" onclick="exportYear()">Xuất Excel năm này</button>
      </div>
      <div id="monthGrid" class="month-grid"></div>
    </div>

    <h2 style="margin-top:24px">Chi tiết theo ngày</h2>
    <div class="hist-top">
      <div class="hist-pick">
        <label>Xem ngày</label>
        <select id="datePick" onchange="loadSales(this.value)"></select>
      </div>
      <div class="stat">
        <div class="stat-lbl">Doanh thu ngày này</div>
        <div class="stat-val num" id="dayRevenue">0 ₫</div>
      </div>
      <div class="stat">
        <div class="stat-lbl">Số đơn ngày này</div>
        <div class="stat-val num" id="dayCount">0</div>
      </div>
      <div class="hist-actions">
        <button class="btn" onclick="exportExcel(false)">Xuất Excel ngày này</button>
        <button class="btn ghost" onclick="exportExcel(true)">Xuất tất cả</button>
      </div>
    </div>
    <div id="salesList" class="sales-list"></div>
  </section>

  <!-- ================= SỔ NỢ ================= -->
  <section id="view-debt" class="hide">
    <div class="debt-search">
      <input id="debtPhone" placeholder="Nhập số điện thoại khách nợ…" inputmode="tel" autocomplete="off"
             onkeydown="if(event.key==='Enter')lookupDebt()">
      <button class="btn" onclick="lookupDebt()">Tra cứu</button>
    </div>
    <div id="debtList"></div>
    <div id="debtResult"></div>
  </section>
</main>

<div id="toast"></div>

<!-- Màn hình đăng nhập -->
<div id="loginScreen" class="login-screen">
  <div class="login-card">
    <h1>Bán hàng</h1>
    <p class="login-sub">Nhập mật khẩu để vào</p>
    <input id="loginPw" type="password" placeholder="Mật khẩu" autocomplete="current-password"
           onkeydown="if(event.key==='Enter')doLogin()">
    <button class="btn" onclick="doLogin()">Đăng nhập</button>
  </div>
</div>

<!-- Modal đổi mật khẩu -->
<div id="changePwModal" class="modal-overlay">
  <div class="modal">
    <h2>Đổi mật khẩu</h2>
    <div class="field"><label>Mật khẩu hiện tại</label>
      <input id="cp-old" type="password" autocomplete="off"></div>
    <div class="field"><label>Mật khẩu mới</label>
      <input id="cp-new" type="password" autocomplete="off"
             onkeydown="if(event.key==='Enter')doChangePw()"></div>
    <div class="modal-actions">
      <button class="btn" onclick="doChangePw()">Lưu</button>
      <button class="btn ghost" onclick="closeChangePw()">Huỷ</button>
    </div>
  </div>
</div>

<!-- Hộp thoại thêm nhanh sản phẩm mới khi đang bán -->
<div id="newProductModal" class="modal-overlay">
  <div class="modal">
    <h2>Sản phẩm mới</h2>
    <div class="np-code-row">Mã vạch: <span id="np-code" class="num"></span></div>
    <input type="hidden" id="np-barcode">
    <div class="field"><label>Tên sản phẩm</label><input id="np-name" autocomplete="off"></div>
    <div class="field"><label>Giá bán (₫)</label>
      <input id="np-price" inputmode="numeric" autocomplete="off"
             onkeydown="if(event.key==='Enter')confirmNewProduct()"></div>
    <div class="field"><label>Tồn kho (tuỳ chọn)</label>
      <input id="np-stock" inputmode="numeric" autocomplete="off"
             onkeydown="if(event.key==='Enter')confirmNewProduct()"></div>
    <div class="modal-actions">
      <button class="btn" onclick="confirmNewProduct()">Thêm & vào giỏ</button>
      <button class="btn ghost" onclick="closeNewProductModal()">Huỷ</button>
    </div>
  </div>
</div>

<script>
const fmt = n => (Number(n)||0).toLocaleString('vi-VN') + ' ₫';
const api = (url, opt={}) => fetch(url, {cache:'no-store', ...opt}).then(r => r.ok ? r.json() : r.json().then(e=>Promise.reject(e)));
function toast(msg){const t=document.getElementById('toast');t.textContent=msg;t.classList.add('show');
  clearTimeout(t._t);t._t=setTimeout(()=>t.classList.remove('show'),1800);}

function showTab(name){
  document.getElementById('view-sale').classList.toggle('hide', name!=='sale');
  document.getElementById('view-prod').classList.toggle('hide', name!=='prod');
  document.getElementById('view-hist').classList.toggle('hide', name!=='hist');
  document.getElementById('view-debt').classList.toggle('hide', name!=='debt');
  document.getElementById('tab-sale').classList.toggle('active', name==='sale');
  document.getElementById('tab-prod').classList.toggle('active', name==='prod');
  document.getElementById('tab-hist').classList.toggle('active', name==='hist');
  document.getElementById('tab-debt').classList.toggle('active', name==='debt');
  if(name==='sale'){ loadSaleList(); document.getElementById('scan').focus(); }
  else if(name==='prod'){ loadProducts(); }
  else if(name==='hist'){ loadSales(); loadYearly(); }
  else if(name==='debt'){ document.getElementById('debtPhone').focus(); loadDebtors(); }
}

/* ---------- BÁN HÀNG ---------- */
let cart = [];   // {id,name,price,qty}

function loadSaleList(){
  api('/api/products').then(list=>{
    const el=document.getElementById('saleList');
    if(!list.length){el.innerHTML='<p style="color:var(--muted)">Chưa có sản phẩm. Sang tab Sản phẩm để thêm.</p>';return;}
    el.innerHTML = list.map(p=>`
      <button class="pcard" onclick='addToCart(${JSON.stringify(p).replace(/'/g,"&#39;")})'>
        <div class="pn">${esc(p.name)}</div>
        <div class="pp num">${fmt(p.price)}</div>
        <div class="ps">Tồn: ${p.stock}</div>
      </button>`).join('');
  });
}

function addToCart(p){
  const f = cart.find(i=>i.id===p.id);
  if(f) f.qty++;
  else cart.push({id:p.id,name:p.name,price:p.price,qty:1});
  renderCart();
}
function changeQty(id,delta){
  const i=cart.find(x=>x.id===id); if(!i)return;
  i.qty+=delta; if(i.qty<=0) cart=cart.filter(x=>x.id!==id);
  renderCart();
}
function removeItem(id){cart=cart.filter(x=>x.id!==id);renderCart();}
function clearCart(){cart=[];document.getElementById('paid').value='';renderCart();}

function renderCart(){
  const box=document.getElementById('cartItems');
  if(!cart.length){
    box.innerHTML='<div class="cempty">Giỏ hàng trống</div>';
  } else {
    box.innerHTML=cart.map(i=>`
      <div class="citem">
        <div class="ci-n"><div>${esc(i.name)}</div>
          <div class="ci-price num" onclick="editCartPrice(${i.id})" title="Bấm để sửa giá">${fmt(i.price)} ✎</div>
        </div>
        <div class="qty">
          <button onclick="changeQty(${i.id},-1)">−</button>
          <span class="num">${i.qty}</span>
          <button onclick="changeQty(${i.id},1)">+</button>
        </div>
        <button class="ci-rm" onclick="removeItem(${i.id})" title="Xoá">×</button>
      </div>`).join('');
  }
  const total=cart.reduce((s,i)=>s+i.price*i.qty,0);
  document.getElementById('total').textContent=fmt(total);
  onPaid();
  document.getElementById('payBtn').disabled = cart.length===0;
  document.getElementById('debtBtn').disabled = cart.length===0;
  focusScan();
}

// đưa con trỏ về ô quét để tiếp tục bán (bỏ qua khi đang mở hộp thoại / ở tab khác)
function focusScan(){
  const el = document.getElementById('scan');
  if(!el) return;
  if(document.getElementById('view-sale').classList.contains('hide')) return;
  if(document.getElementById('loginScreen').classList.contains('show')) return;
  if(document.getElementById('newProductModal').classList.contains('show')) return;
  if(document.getElementById('changePwModal').classList.contains('show')) return;
  el.focus();
}

function onPaid(){
  const total=cart.reduce((s,i)=>s+i.price*i.qty,0);
  const paid=parseInt((document.getElementById('paid').value||'').replace(/\D/g,''))||0;
  const change=paid-total;
  const el=document.getElementById('change');
  el.textContent = change>=0 ? fmt(change) : '—';
  el.style.color = change>=0 ? 'var(--accent)' : 'var(--danger)';
}

function editCartPrice(id){
  const item = cart.find(x=>x.id===id);
  if(!item) return;
  const input = prompt('Giá bán mới cho "'+item.name+'":', item.price);
  if(input===null) return;                       // bấm Huỷ
  const newPrice = parseInt((input||'').replace(/\D/g,''))||0;
  if(newPrice<=0){ toast('Giá không hợp lệ'); return; }
  item.price = newPrice;                          // cập nhật đơn đang bán
  renderCart();
  // lưu giá mới xuống kho để lần sau bán dùng giá này
  api('/api/products/'+id+'/price',{method:'POST',
    headers:{'Content-Type':'application/json'},body:JSON.stringify({price:newPrice})})
    .then(()=>{ toast('Đã cập nhật giá'); loadSaleList(); })
    .catch(()=> toast('Đã đổi giá cho đơn này (chưa lưu được vào kho)'));
}

function checkout(){
  const total=cart.reduce((s,i)=>s+i.price*i.qty,0);
  const paid=parseInt((document.getElementById('paid').value||'').replace(/\D/g,''))||0;
  if(paid && paid<total){toast('Khách đưa chưa đủ tiền');return;}
  api('/api/checkout',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({items:cart,paid:paid||total})})
    .then(r=>{
      toast(r.change>0 ? 'Đã thanh toán · Thối '+fmt(r.change) : 'Đã thanh toán');
      clearCart(); loadSaleList();
    }).catch(e=>toast(e.error||'Lỗi thanh toán'));
}

function checkoutDebt(){
  const total=cart.reduce((s,i)=>s+i.price*i.qty,0);
  const phone=(prompt('Nhập số điện thoại khách nợ ('+fmt(total)+'):')||'').trim();
  if(!phone) return;
  api('/api/checkout',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({items:cart,debtor_phone:phone})})
    .then(()=>{
      toast('Đã ghi nợ '+fmt(total)+' cho '+phone);
      clearCart(); loadSaleList();
    }).catch(e=>toast(e.error||'Lỗi ghi nợ'));
}

// quét mã vạch: tự tìm sản phẩm và thêm vào giỏ, KHÔNG cần Enter
let scanTimer = null;
let lastScanCode = '';
let lastScanAt = 0;

function processScan(){
  const el = document.getElementById('scan');
  const code = (el.value || '').trim();
  el.value = '';                 // xoá ngay để tránh xử lý trùng
  if(!code) return;
  // chống quét đúp: cùng một mã trong vòng 400ms chỉ tính 1 lần
  const now = Date.now();
  if(code === lastScanCode && (now - lastScanAt) < 400) return;
  lastScanCode = code;
  lastScanAt = now;
  api('/api/products/barcode/' + encodeURIComponent(code))
    .then(p=>{ addToCart(p); toast('Đã thêm ' + p.name); el.focus(); })
    .catch(()=>{
      // chưa có trong kho -> tra Open Food Facts rồi mở hộp thoại nhập giá
      toast('Mã mới — đang tra thông tin…');
      api('/api/lookup/' + encodeURIComponent(code))
        .then(d=> openNewProductModal(code, d.found ? d.name : ''))
        .catch(()=> openNewProductModal(code, ''));
    });
}

// hộp thoại thêm nhanh sản phẩm mới ngay khi đang bán
function openNewProductModal(code, name){
  document.getElementById('np-code').textContent = code;
  document.getElementById('np-barcode').value = code;
  document.getElementById('np-name').value = name || '';
  document.getElementById('np-price').value = '';
  document.getElementById('np-stock').value = '';
  document.getElementById('newProductModal').classList.add('show');
  setTimeout(()=>{
    (name ? document.getElementById('np-price') : document.getElementById('np-name')).focus();
  }, 50);
}

function confirmNewProduct(){
  const name = document.getElementById('np-name').value.trim();
  const barcode = document.getElementById('np-barcode').value.trim();
  const price = parseInt((document.getElementById('np-price').value||'').replace(/\D/g,''))||0;
  const stock = parseInt((document.getElementById('np-stock').value||'').replace(/\D/g,''))||0;
  if(!name){ toast('Nhập tên sản phẩm'); document.getElementById('np-name').focus(); return; }
  if(price<=0){ toast('Nhập giá bán'); document.getElementById('np-price').focus(); return; }
  api('/api/products',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({name, barcode, price, stock})})
    .then(p=>{
      addToCart(p);                 // luồng giỏ hàng
      toast('Đã thêm "'+p.name+'" vào kho & giỏ');
      closeNewProductModal();
      loadSaleList();               // luồng cập nhật danh sách sản phẩm
    })
    .catch(e=> toast(e.error||'Lỗi thêm sản phẩm'));
}

function closeNewProductModal(){
  document.getElementById('newProductModal').classList.remove('show');
  document.getElementById('scan').focus();
}

document.addEventListener('DOMContentLoaded',()=>{
  const el = document.getElementById('scan');
  // nếu máy quét có gửi Enter -> xử lý ngay lập tức
  el.addEventListener('keydown', e=>{
    if(e.isComposing || e.keyCode === 229) return;   // bộ gõ tiếng Việt đang ghép ký tự -> bỏ qua
    if(e.key === 'Enter'){ e.preventDefault(); clearTimeout(scanTimer); processScan(); }
  });
  // nếu máy quét KHÔNG gửi Enter -> tự xử lý khi dòng ký tự ngừng lại
  el.addEventListener('input', e=>{
    if(e.isComposing) return;                          // bỏ qua sự kiện khi bộ gõ đang ghép
    clearTimeout(scanTimer);
    scanTimer = setTimeout(()=>{
      if((el.value || '').trim().length >= 6) processScan();   // đủ dài mới tự thêm (mã vạch thường 8–13 số)
    }, 120);
  });

  // ô Mã vạch trong form Sản phẩm: cũng nhận máy quét
  const bc = document.getElementById('f-barcode');
  bc.addEventListener('keydown', e=>{
    if(e.isComposing || e.keyCode === 229) return;   // bộ gõ tiếng Việt đang ghép -> bỏ qua
    if(e.key === 'Enter'){ e.preventDefault(); clearTimeout(barcodeTimer); processFormBarcode(); }
  });
  bc.addEventListener('input', e=>{
    if(e.isComposing) return;
    clearTimeout(barcodeTimer);
    barcodeTimer = setTimeout(()=>{
      if((bc.value || '').trim().length >= 6) processFormBarcode();
    }, 120);
  });
});

// xử lý mã vạch quét vào form Sản phẩm
let barcodeTimer = null;
let lastBarcode = '';
function processFormBarcode(){
  const code = (document.getElementById('f-barcode').value || '').trim();
  if(!code || code === lastBarcode) return;   // tránh xử lý lặp cùng một mã
  lastBarcode = code;
  api('/api/products/barcode/' + encodeURIComponent(code))
    .then(p=>{ editProduct(p); toast('Mã này đã có — đang sửa "' + p.name + '"'); })
    .catch(()=>{
      // chưa có trong kho -> hỏi Open Food Facts để gợi ý tên
      toast('Mã mới — đang tra thông tin…');
      api('/api/lookup/' + encodeURIComponent(code))
        .then(d=>{
          if(d.found && d.name){
            document.getElementById('f-name').value = d.name;
            toast('Đã điền tên: ' + d.name + ' — nhập giá & tồn');
            document.getElementById('f-price').focus();
          } else {
            toast('Không có sẵn thông tin — nhập tên tay');
            document.getElementById('f-name').focus();
          }
        })
        .catch(()=>{ document.getElementById('f-name').focus(); });
    });
}

/* ---------- SẢN PHẨM ---------- */
function loadProducts(q=''){
  api('/api/products'+(q?('?q='+encodeURIComponent(q)):'')).then(list=>{
    const tb=document.getElementById('prodTable');
    if(!list.length){tb.innerHTML='<tr class="empty-row"><td colspan="5">Chưa có sản phẩm nào</td></tr>';return;}
    tb.innerHTML=list.map(p=>`
      <tr>
        <td>${esc(p.name)}</td>
        <td class="num">${esc(p.barcode||'—')}</td>
        <td class="r num">${fmt(p.price)}</td>
        <td class="r num">${p.stock}</td>
        <td class="r t-actions">
          <button onclick='editProduct(${JSON.stringify(p).replace(/'/g,"&#39;")})'>Sửa</button>
          <button class="del" onclick="deleteProduct(${p.id},'${esc(p.name)}')">Xoá</button>
        </td>
      </tr>`).join('');
  });
}

function saveProduct(){
  const id=document.getElementById('pid').value;
  const body={
    name:document.getElementById('f-name').value,
    barcode:document.getElementById('f-barcode').value,
    price:parseInt((document.getElementById('f-price').value||'').replace(/\D/g,''))||0,
    stock:parseInt((document.getElementById('f-stock').value||'').replace(/\D/g,''))||0,
  };
  if(!body.name.trim()){toast('Nhập tên sản phẩm');return;}
  const opt={method:id?'PUT':'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)};
  api(id?('/api/products/'+id):'/api/products',opt)
    .then(()=>{toast(id?'Đã cập nhật':'Đã thêm sản phẩm');resetForm();loadProducts(document.getElementById('search').value);})
    .catch(e=>toast(e.error||'Lỗi lưu'));
}

function editProduct(p){
  document.getElementById('pid').value=p.id;
  document.getElementById('f-name').value=p.name;
  document.getElementById('f-barcode').value=p.barcode||'';
  document.getElementById('f-price').value=p.price;
  document.getElementById('f-stock').value=p.stock;
  document.getElementById('formTitle').textContent='Sửa sản phẩm';
  document.getElementById('cancelEdit').style.display='';
  document.getElementById('f-name').focus();
}

function deleteProduct(id,name){
  if(!confirm('Xoá "'+name+'"?'))return;
  api('/api/products/'+id,{method:'DELETE'}).then(()=>{toast('Đã xoá');loadProducts(document.getElementById('search').value);});
}

function resetForm(){
  document.getElementById('pid').value='';
  ['f-name','f-barcode','f-price','f-stock'].forEach(i=>document.getElementById(i).value='');
  document.getElementById('formTitle').textContent='Thêm sản phẩm';
  document.getElementById('cancelEdit').style.display='none';
  lastBarcode='';
}

function esc(s){return String(s).replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}

/* ---------- LỊCH SỬ ---------- */
let currentDate = '';

function loadSales(date){
  const url = '/api/sales' + (date ? ('?date=' + encodeURIComponent(date)) : '');
  api(url).then(d=>{
    currentDate = d.date;
    document.getElementById('dayRevenue').textContent = fmt(d.day_revenue);
    document.getElementById('dayCount').textContent = d.day_count;

    // đổ danh sách ngày vào dropdown
    const sel = document.getElementById('datePick');
    const opts = (d.dates && d.dates.length) ? d.dates : [d.date];
    if(!opts.includes(d.date)) opts.unshift(d.date);
    sel.innerHTML = opts.map(x=>`<option value="${x}" ${x===d.date?'selected':''}>${x}</option>`).join('');

    const box = document.getElementById('salesList');
    if(!d.sales.length){
      box.innerHTML='<div class="cempty">Không có giao dịch trong ngày này</div>';
      return;
    }
    box.innerHTML = d.sales.map(s=>{
      const items = s.items.map(i=>
        `<div class="bill-item"><span>${esc(i.name)} ×${i.qty}</span><span class="num">${fmt(i.price*i.qty)}</span></div>`
      ).join('');
      return `
        <div class="bill">
          <div class="bill-head">
            <span class="bill-id">#${s.id}</span>
            <span class="bill-time">${esc(s.created_at)}</span>
          </div>
          <div class="bill-body">${items}</div>
          <div class="bill-foot">
            <div><span class="lbl">Tổng</span> <span class="num bill-total">${fmt(s.total)}</span></div>
            <div class="bill-sub"><span class="lbl">Khách đưa</span> <span class="num">${fmt(s.paid)}</span> ·
                 <span class="lbl">Thối</span> <span class="num">${fmt(s.change)}</span></div>
          </div>
        </div>`;
    }).join('');
  });
}

function exportExcel(all){
  const url = all ? '/api/export?date=all'
                  : ('/api/export?date=' + encodeURIComponent(currentDate));
  window.location.href = url;   // trình duyệt tải file Excel về
}

/* ---------- DOANH THU THEO NĂM ---------- */
let currentYear = '';
const MONTH_NAMES = ['Th1','Th2','Th3','Th4','Th5','Th6','Th7','Th8','Th9','Th10','Th11','Th12'];

function loadYearly(year){
  const url = '/api/revenue/yearly' + (year ? ('?year=' + encodeURIComponent(year)) : '');
  api(url).then(d=>{
    currentYear = d.year;
    document.getElementById('yearTotal').textContent = fmt(d.total);
    document.getElementById('yearCount').textContent = d.count;

    const sel = document.getElementById('yearPick');
    const opts = (d.years && d.years.length) ? d.years.slice() : [d.year];
    if(!opts.includes(d.year)) opts.unshift(d.year);
    sel.innerHTML = opts.map(y=>`<option value="${y}" ${y===d.year?'selected':''}>${y}</option>`).join('');

    const max = Math.max(1, ...d.months.map(m=>m.revenue));
    document.getElementById('monthGrid').innerHTML = d.months.map((m,idx)=>{
      const pct = Math.round(m.revenue / max * 100);
      return `<div class="month-cell">
        <div class="month-name">${MONTH_NAMES[idx]}</div>
        <div class="month-bar"><div class="month-fill" style="height:${pct}%"></div></div>
        <div class="month-rev num">${m.revenue ? fmt(m.revenue) : '—'}</div>
        <div class="month-cnt num">${m.count} đơn</div>
      </div>`;
    }).join('');
  });
}

function exportYear(){
  window.location.href = '/api/export?year=' + encodeURIComponent(currentYear);
}

/* ---------- SỔ NỢ ---------- */
let debtPhone = '';

function loadDebtors(){
  api('/api/debtors').then(list=>{
    const box=document.getElementById('debtList');
    if(!list.length){
      box.innerHTML='<div class="cempty">Hiện không có ai đang nợ</div>';
      return;
    }
    box.innerHTML='<h2>Đang nợ ('+list.length+')</h2>'+
      '<div class="debtor-list">'+
      list.map((d,i)=>`
        <button class="debtor-row" onclick="openDebtor('${esc(d.phone)}')">
          <span class="debtor-rank">${i+1}</span>
          <span class="debtor-phone">${esc(d.phone)}</span>
          <span class="debtor-amt num">${fmt(d.remaining)}</span>
        </button>`).join('')+
      '</div>';
  });
}

function openDebtor(phone){
  document.getElementById('debtPhone').value=phone;
  lookupDebt();
  document.getElementById('debtResult').scrollIntoView({behavior:'smooth',block:'start'});
}

function lookupDebt(){
  const phone=(document.getElementById('debtPhone').value||'').trim();
  if(!phone){toast('Nhập số điện thoại');return;}
  debtPhone = phone;
  api('/api/debt/'+encodeURIComponent(phone)).then(d=>{
    const box=document.getElementById('debtResult');
    if(!d.orders.length){
      box.innerHTML='<div class="cempty">Khách '+esc(phone)+' không có khoản nợ nào</div>';
      return;
    }
    const orders = d.orders.map(s=>{
      const items=s.items.map(i=>
        `<div class="bill-item"><span>${esc(i.name)} ×${i.qty}</span><span class="num">${fmt(i.price*i.qty)}</span></div>`
      ).join('');
      return `<div class="bill">
        <div class="bill-head"><span class="bill-id">#${s.id}</span><span class="bill-time">${esc(s.created_at)}</span></div>
        <div class="bill-body">${items}</div>
        <div class="bill-foot"><div><span class="lbl">Tiền đơn</span> <span class="num bill-total">${fmt(s.total)}</span></div></div>
      </div>`;
    }).join('');
    const pays = d.payments.length
      ? d.payments.map(p=>`<div class="bill-item"><span>${esc(p.created_at)}</span><span class="num">− ${fmt(p.amount)}</span></div>`).join('')
      : '<div class="lbl" style="padding:4px 0">Chưa trả lần nào</div>';
    const cleared = d.remaining<=0;
    box.innerHTML=`
      <div class="debt-summary">
        <div>
          <div class="stat-lbl">Còn nợ</div>
          <div class="stat-val num" style="color:${cleared?'var(--accent)':'var(--danger)'}">${cleared?'Đã trả hết':fmt(d.remaining)}</div>
          <div class="lbl num" style="margin-top:4px">Đã mua ${fmt(d.debt_total)} · đã trả ${fmt(d.paid_total)}</div>
        </div>
        <div class="debt-pay">
          <input id="payAmount" class="field-inline num" inputmode="numeric" placeholder="Số tiền khách trả"
                 onkeydown="if(event.key==='Enter')payDebt()">
          <button class="btn" onclick="payDebt()">Ghi nhận trả nợ</button>
          <button class="btn btn-red" onclick="closeDebtDetail()">Xong</button>
        </div>
      </div>
      <div class="debt-cols">
        <div><h2>Các đơn đã nợ</h2>${orders}</div>
        <div><h2>Lịch sử trả nợ</h2><div class="pay-history">${pays}</div></div>
      </div>`;
  }).catch(()=>toast('Lỗi tra cứu'));
}

function payDebt(){
  const amount=parseInt((document.getElementById('payAmount').value||'').replace(/\D/g,''))||0;
  if(amount<=0){toast('Nhập số tiền trả');return;}
  api('/api/debt/'+encodeURIComponent(debtPhone)+'/pay',{method:'POST',
    headers:{'Content-Type':'application/json'},body:JSON.stringify({amount})})
    .then(r=>{
      toast(r.remaining<=0 ? 'Đã trả hết nợ' : 'Đã ghi nhận · Còn nợ '+fmt(r.remaining));
      lookupDebt(); loadDebtors();
    }).catch(e=>toast(e.error||'Lỗi'));
}

function closeDebtDetail(){
  document.getElementById('debtResult').innerHTML='';
  document.getElementById('debtPhone').value='';
  loadDebtors();
  window.scrollTo({top:0, behavior:'smooth'});
  document.getElementById('debtPhone').focus();
}

/* ---------- ĐĂNG NHẬP ---------- */
function boot(){
  api('/api/auth/status').then(d=>{
    if(d.logged_in) showApp(); else showLogin();
  }).catch(()=> showLogin());
}
function showLogin(){
  document.getElementById('loginScreen').classList.add('show');
  setTimeout(()=>document.getElementById('loginPw').focus(), 50);
}
function showApp(){
  document.getElementById('loginScreen').classList.remove('show');
  renderCart(); loadSaleList();
  document.getElementById('scan').focus();
}
function doLogin(){
  const pw = document.getElementById('loginPw').value;
  api('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({password:pw})})
    .then(()=>{ document.getElementById('loginPw').value=''; showApp(); })
    .catch(()=>{ toast('Sai mật khẩu'); document.getElementById('loginPw').value=''; document.getElementById('loginPw').focus(); });
}
function doLogout(){
  api('/api/logout',{method:'POST'}).then(()=>{ location.reload(); });
}
function openChangePw(){
  document.getElementById('cp-old').value='';
  document.getElementById('cp-new').value='';
  document.getElementById('changePwModal').classList.add('show');
  setTimeout(()=>document.getElementById('cp-old').focus(), 50);
}
function closeChangePw(){ document.getElementById('changePwModal').classList.remove('show'); }
function doChangePw(){
  const oldp=document.getElementById('cp-old').value;
  const newp=document.getElementById('cp-new').value;
  api('/api/change-password',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({old:oldp, new:newp})})
    .then(()=>{ toast('Đã đổi mật khẩu'); closeChangePw(); })
    .catch(e=> toast(e.error||'Lỗi đổi mật khẩu'));
}

// khởi động
boot();
</script>
</body>
</html>
"""


@app.before_request
def require_login():
    """Chặn mọi API nếu chưa đăng nhập (trừ các route đăng nhập/trạng thái)."""
    p = request.path
    if p == "/" or p.startswith("/static"):
        return
    if p in ("/api/login", "/api/logout", "/api/auth/status"):
        return
    if p.startswith("/api/") and not session.get("auth"):
        return jsonify({"error": "unauthorized"}), 401


@app.after_request
def no_cache_api(resp):
    """Không cho trình duyệt cache các phản hồi API (tránh thấy trạng thái cũ sau đăng xuất)."""
    if request.path.startswith("/api/"):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/api/auth/status", methods=["GET"])
def auth_status():
    return jsonify({"logged_in": bool(session.get("auth"))})


@app.route("/api/login", methods=["POST"])
def login():
    d = request.get_json(force=True)
    pw = d.get("password") or ""
    if check_password_hash(get_setting("password_hash") or "", pw):
        session["auth"] = True
        return jsonify({"ok": True})
    return jsonify({"error": "Sai mật khẩu"}), 401


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/change-password", methods=["POST"])
def change_password():
    d = request.get_json(force=True)
    old = d.get("old") or ""
    new = (d.get("new") or "").strip()
    if not check_password_hash(get_setting("password_hash") or "", old):
        return jsonify({"error": "Mật khẩu hiện tại không đúng"}), 400
    if len(new) < 4:
        return jsonify({"error": "Mật khẩu mới phải từ 4 ký tự"}), 400
    set_setting("password_hash", generate_password_hash(new))
    return jsonify({"ok": True})


@app.route("/")
def index():
    return render_template_string(PAGE)


# ---------------------------------------------------------------------------
# Khởi chạy: Flask chạy nền, PyWebView mở cửa sổ app
# ---------------------------------------------------------------------------
# Cổng cho bản desktop. Dùng 8770 (không phải 5000) vì macOS chiếm 5000 cho AirPlay.
APP_PORT = int(os.environ.get("PORT", "8770"))


def run_flask():
    host = os.environ.get("HOST", "127.0.0.1")
    app.run(host=host, port=APP_PORT, threaded=True)


# Khởi tạo DB ngay khi nạp module (cần cho gunicorn trong Docker, vốn không chạy __main__)
init_db()
app.secret_key = get_setting("secret_key") or os.urandom(24).hex()

if __name__ == "__main__":
    url = f"http://127.0.0.1:{APP_PORT}"
    try:
        import webview  # cửa sổ desktop
        threading.Thread(target=run_flask, daemon=True).start()
        webview.create_window("Bán hàng", url,
                              width=1180, height=760, min_size=(900, 600))
        webview.start()
    except ImportError:
        # Chưa cài pywebview -> chạy như web thường, mở trình duyệt vào địa chỉ này
        print("Chưa có pywebview. Mở trình duyệt: " + url)
        run_flask()
